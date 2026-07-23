"""
Excel export for expense records (openpyxl).
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.expenses.choices import ExpenseStatus

HEADER_FILL = PatternFill("solid", fgColor="0F766E")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)

HEADERS = [
    "Id",
    "Title",
    "Employee",
    "Role",
    "Team",
    "Type",
    "Status",
    "Claimed Amount",
    "Approved Amount",
    "Expense Date",
    "Project / Lead",
    "Comment",
    "Reviewed By",
    "Reviewed At",
    "Receipt",
]


def _auto_width(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
            length + 2,
            40,
        )


def _display_id(expense) -> str:
    short = str(expense.id).replace("-", "")[:6].upper()
    return f"EXP-{short}"


def _title(expense) -> str:
    description = (expense.description or "").strip()
    if description:
        return description[:80]
    return f"{expense.get_category_display()} Expense"


def _employee_name(expense) -> str:
    return expense.submitted_by.get_full_name() or expense.submitted_by.username


def _employee_role(expense) -> str:
    user = expense.submitted_by
    if hasattr(user, "get_role_display"):
        return user.get_role_display()
    return user.role


def _team_name(expense) -> str:
    manager = getattr(expense.submitted_by, "manager", None)
    if manager is None:
        return "NA"
    return manager.get_full_name() or manager.username


def _lead_name(expense) -> str:
    if not expense.lead_id:
        return "NA"
    return expense.lead.customer_name or expense.lead.company_name or "NA"


def _approved_amount(expense):
    if expense.status == ExpenseStatus.APPROVED:
        return float(expense.amount)
    return "NA"


def build_expenses_workbook(expenses, *, title: str = "All Expenses") -> BytesIO:
    """Build an Excel workbook for the given expense queryset/list."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Total rows: {len(expenses)}")

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_index, expense in enumerate(expenses, start=5):
        reviewed_by = ""
        if expense.reviewed_by_id:
            reviewed_by = (
                expense.reviewed_by.get_full_name() or expense.reviewed_by.username
            )
        reviewed_at = ""
        if expense.reviewed_at:
            reviewed_at = expense.reviewed_at.strftime("%Y-%m-%d %H:%M")

        values = [
            _display_id(expense),
            _title(expense),
            _employee_name(expense),
            _employee_role(expense),
            _team_name(expense),
            f"{expense.get_category_display()} Expense",
            expense.get_status_display(),
            float(expense.amount),
            _approved_amount(expense),
            expense.expense_date.isoformat() if expense.expense_date else "",
            _lead_name(expense),
            (expense.review_notes or "").strip() or "NA",
            reviewed_by or "NA",
            reviewed_at or "NA",
            "Yes" if expense.receipt else "No",
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row_index, column=col, value=value)

    _auto_width(ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
