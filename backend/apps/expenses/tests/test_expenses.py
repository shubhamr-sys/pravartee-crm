"""
Expense workflow tests.
"""
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.utils import timezone
from rest_framework import status
from rest_framework.test import APIClient

from apps.accounts.choices import UserRole
from apps.activities.models import ActivityType, LeadActivity
from apps.expenses.choices import ExpenseCategory, ExpenseStatus
from apps.expenses.models import Expense
from apps.leads.models import Lead, LeadStage, ProductCategory

User = get_user_model()


class ExpenseWorkflowTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.category = ProductCategory.objects.get(name="IT")
        cls.stage = LeadStage.objects.get(name="New")

        cls.ceo = User.objects.create_user(
            username="ceo_expense",
            email="ceo_expense@test.com",
            password="pass12345",
            role=UserRole.CEO,
        )
        cls.sales_head = User.objects.create_user(
            username="head_expense",
            email="head_expense@test.com",
            password="pass12345",
            role=UserRole.SALES_HEAD,
        )
        cls.salesperson = User.objects.create_user(
            username="sales_expense",
            email="sales_expense@test.com",
            password="pass12345",
            role=UserRole.SALESPERSON,
        )
        cls.other_sales = User.objects.create_user(
            username="other_expense",
            email="other_expense@test.com",
            password="pass12345",
            role=UserRole.SALESPERSON,
        )
        cls.accounts = User.objects.create_user(
            username="accounts_expense",
            email="accounts_expense@test.com",
            password="pass12345",
            role=UserRole.ACCOUNTS,
        )

        cls.sales_head.manager = cls.ceo
        cls.sales_head.save(update_fields=["manager"])
        cls.salesperson.manager = cls.sales_head
        cls.salesperson.save(update_fields=["manager"])
        cls.other_sales.manager = cls.sales_head
        cls.other_sales.save(update_fields=["manager"])

        cls.lead = Lead.objects.create(
            customer_name="Expense Lead",
            company_name="Expense Co",
            category=cls.category,
            stage=cls.stage,
            assigned_to=cls.salesperson,
        )

    def setUp(self):
        self.client = APIClient()

    def _receipt_file(self, name="receipt.pdf"):
        return SimpleUploadedFile(
            name,
            b"%PDF-1.4 test receipt content",
            content_type="application/pdf",
        )

    def _create_expense(self, user, **kwargs):
        if "receipt" not in kwargs:
            kwargs["receipt"] = self._receipt_file()
        return Expense.objects.create(
            submitted_by=user,
            category=ExpenseCategory.TRAVEL,
            amount=Decimal("1500.00"),
            expense_date=timezone.localdate(),
            description="Client visit travel",
            lead=self.lead,
            **kwargs,
        )

    def test_salesperson_can_create_expense(self):
        self.client.force_authenticate(user=self.salesperson)
        response = self.client.post(
            "/api/v1/expenses/",
            {
                "category": ExpenseCategory.FOOD,
                "amount": "450.00",
                "expense_date": "2026-07-01",
                "description": "Lunch with client",
                "lead": str(self.lead.id),
                "receipt": self._receipt_file(),
            },
            format="multipart",
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["status"], ExpenseStatus.PENDING)
        self.assertEqual(Expense.objects.filter(submitted_by=self.salesperson).count(), 1)

    def test_create_expense_requires_receipt(self):
        self.client.force_authenticate(user=self.salesperson)
        response = self.client.post(
            "/api/v1/expenses/",
            {
                "category": ExpenseCategory.FOOD,
                "amount": "450.00",
                "expense_date": "2026-07-01",
                "description": "Lunch with client",
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("receipt", response.data)

    def test_create_expense_requires_category_amount_and_date(self):
        self.client.force_authenticate(user=self.salesperson)
        response = self.client.post(
            "/api/v1/expenses/",
            {
                "description": "Missing required fields",
                "receipt": self._receipt_file(),
            },
            format="multipart",
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("category", response.data)
        self.assertIn("amount", response.data)
        self.assertIn("expense_date", response.data)

    def test_salesperson_sees_only_own_expenses(self):
        self._create_expense(self.salesperson)
        self._create_expense(self.other_sales)
        self.client.force_authenticate(user=self.salesperson)
        response = self.client.get("/api/v1/expenses/", {"tab": "all"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data["results"]), 1)

    def test_ceo_can_see_team_expenses(self):
        self._create_expense(self.salesperson)
        self._create_expense(self.other_sales)
        self.client.force_authenticate(user=self.ceo)
        response = self.client.get("/api/v1/expenses/", {"tab": "all"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data["results"]), 2)

    def test_pending_tab_filters_pending_expenses(self):
        pending = self._create_expense(self.salesperson)
        approved = self._create_expense(
            self.other_sales,
            status=ExpenseStatus.APPROVED,
            reviewed_by=self.ceo,
            reviewed_at=timezone.now(),
        )
        self.client.force_authenticate(user=self.ceo)
        response = self.client.get("/api/v1/expenses/", {"tab": "pending"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        ids = {item["id"] for item in response.data["results"]}
        self.assertIn(str(pending.id), ids)
        self.assertNotIn(str(approved.id), ids)

    def test_sales_head_cannot_approve_expense(self):
        expense = self._create_expense(self.salesperson)
        self.client.force_authenticate(user=self.sales_head)
        response = self.client.post(
            f"/api/v1/expenses/{expense.id}/approve/",
            {"review_notes": "Approved for reimbursement"},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        expense.refresh_from_db()
        self.assertEqual(expense.status, ExpenseStatus.PENDING)

    def test_ceo_can_approve_expense(self):
        expense = self._create_expense(self.salesperson)
        self.client.force_authenticate(user=self.ceo)
        response = self.client.post(
            f"/api/v1/expenses/{expense.id}/approve/",
            {"review_notes": "Approved for reimbursement"},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        expense.refresh_from_db()
        self.assertEqual(expense.status, ExpenseStatus.APPROVED)
        self.assertEqual(expense.reviewed_by_id, self.ceo.id)

    def test_salesperson_cannot_approve_own_expense(self):
        expense = self._create_expense(self.salesperson)
        self.client.force_authenticate(user=self.salesperson)
        response = self.client.post(f"/api/v1/expenses/{expense.id}/approve/")
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_expense_summary(self):
        self._create_expense(self.salesperson)
        self._create_expense(
            self.other_sales,
            status=ExpenseStatus.APPROVED,
            reviewed_by=self.ceo,
            reviewed_at=timezone.now(),
        )
        self.client.force_authenticate(user=self.ceo)
        response = self.client.get("/api/v1/expenses/summary/")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["pending_count"], 1)
        self.assertEqual(response.data["approved_count"], 1)

    def test_accounts_user_sees_all_expenses(self):
        self._create_expense(self.salesperson)
        self._create_expense(self.other_sales)
        self.client.force_authenticate(user=self.accounts)
        response = self.client.get("/api/v1/expenses/", {"tab": "all"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data["results"]), 2)

    def test_accounts_user_can_approve_expense(self):
        expense = self._create_expense(self.salesperson)
        self.client.force_authenticate(user=self.accounts)
        response = self.client.post(
            f"/api/v1/expenses/{expense.id}/approve/",
            {"review_notes": "Approved by accounts"},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        expense.refresh_from_db()
        self.assertEqual(expense.status, ExpenseStatus.APPROVED)
        self.assertEqual(expense.reviewed_by_id, self.accounts.id)

    def test_accounts_user_cannot_create_expense(self):
        self.client.force_authenticate(user=self.accounts)
        response = self.client.post(
            "/api/v1/expenses/",
            {
                "category": ExpenseCategory.FOOD,
                "amount": "450.00",
                "expense_date": "2026-07-01",
                "description": "Lunch",
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_expense_dashboard_for_accounts(self):
        self._create_expense(self.salesperson)
        self._create_expense(
            self.other_sales,
            status=ExpenseStatus.APPROVED,
            reviewed_by=self.accounts,
            reviewed_at=timezone.now(),
        )
        self.client.force_authenticate(user=self.accounts)
        response = self.client.get("/api/v1/expenses/dashboard/")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["summary"]["pending_count"], 1)
        self.assertEqual(response.data["summary"]["approved_count"], 1)
        self.assertEqual(len(response.data["by_category"]), 5)
        self.assertEqual(len(response.data["recent_pending"]), 1)

    def test_salesperson_cannot_access_expense_dashboard(self):
        self.client.force_authenticate(user=self.salesperson)
        response = self.client.get("/api/v1/expenses/dashboard/")
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_accounts_can_export_expenses_excel(self):
        self._create_expense(self.salesperson)
        self._create_expense(self.other_sales)
        self.client.force_authenticate(user=self.accounts)
        response = self.client.get("/api/v1/expenses/export/", {"tab": "all"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn(
            "spreadsheetml.sheet",
            response["Content-Type"],
        )
        self.assertIn("Expenses_", response["Content-Disposition"])
        content = b"".join(response.streaming_content)
        self.assertGreater(len(content), 0)

    def test_salesperson_export_only_own_expenses(self):
        self._create_expense(self.salesperson)
        self._create_expense(self.other_sales)
        self.client.force_authenticate(user=self.salesperson)
        response = self.client.get("/api/v1/expenses/export/", {"tab": "all"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        from io import BytesIO

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(b"".join(response.streaming_content)))
        ws = wb.active
        self.assertEqual(ws["A2"].value, "Total rows: 1")

    def test_create_expense_with_lead_logs_activity(self):
        self.client.force_authenticate(user=self.salesperson)
        response = self.client.post(
            "/api/v1/expenses/",
            {
                "category": ExpenseCategory.TRAVEL,
                "amount": "900.00",
                "expense_date": "2026-07-10",
                "description": "Site visit travel",
                "lead": str(self.lead.id),
                "receipt": self._receipt_file(),
            },
            format="multipart",
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertTrue(
            LeadActivity.objects.filter(
                lead=self.lead,
                activity_type=ActivityType.EXPENSE_SUBMITTED,
                user=self.salesperson,
            ).exists(),
        )

    def test_approve_expense_with_lead_logs_activity(self):
        expense = self._create_expense(self.salesperson)
        self.client.force_authenticate(user=self.ceo)
        response = self.client.post(
            f"/api/v1/expenses/{expense.id}/approve/",
            {"review_notes": "OK"},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(
            LeadActivity.objects.filter(
                lead=self.lead,
                activity_type=ActivityType.EXPENSE_APPROVED,
                user=self.ceo,
            ).exists(),
        )

    def test_reject_expense_with_lead_logs_activity(self):
        expense = self._create_expense(self.salesperson)
        self.client.force_authenticate(user=self.accounts)
        response = self.client.post(
            f"/api/v1/expenses/{expense.id}/reject/",
            {"review_notes": "Missing receipt"},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(
            LeadActivity.objects.filter(
                lead=self.lead,
                activity_type=ActivityType.EXPENSE_REJECTED,
                user=self.accounts,
            ).exists(),
        )

    def test_expense_without_lead_does_not_log_activity(self):
        self.client.force_authenticate(user=self.salesperson)
        before = LeadActivity.objects.count()
        response = self.client.post(
            "/api/v1/expenses/",
            {
                "category": ExpenseCategory.FOOD,
                "amount": "200.00",
                "expense_date": "2026-07-11",
                "description": "No lead linked",
                "receipt": self._receipt_file(),
            },
            format="multipart",
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(LeadActivity.objects.count(), before)
