"""
Excel export for Sales MBR reports (openpyxl).
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="0F766E")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=12)


def _auto_width(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
            length + 2,
            40,
        )


def _write_section_title(ws, row: int, title: str) -> int:
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    return row + 1


def _write_table(
    ws,
    row: int,
    headers: list[str],
    rows: list[list],
) -> int:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    row += 1
    for data_row in rows:
        for col, value in enumerate(data_row, start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1
    return row + 1


def build_sales_mbr_workbook(report: dict) -> BytesIO:
    """Generate an Excel workbook matching the Sales MBR sheet layout."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales MBR"

    filters = report["filters"]
    summary = report["performance_summary"]
    title = f"Sales MBR — {filters['month_name']} {filters['year']}"

    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    row = 3

    scopes = report.get("metric_scopes") or {}
    if scopes:
        ws.cell(
            row=row,
            column=1,
            value=(
                f"Period: {scopes.get('period', '')} · "
                f"Snapshot: {scopes.get('snapshot', '')}"
            ),
        )
        row += 2

    row = _write_section_title(ws, row, "1. Sales Performance Summary (₹)")
    perf = report.get("sales_performance") or {}
    trading = perf.get("trading") or {}
    solutions = perf.get("solutions") or {}
    total = perf.get("total") or {}
    perf_headers = [
        "Metric",
        "Trading",
        "Solutions",
        "Total (Month)",
        "Target (Month)",
        "Var %",
    ]
    perf_rows = [
        [
            "Order Booking",
            trading.get("order_booking", 0),
            solutions.get("order_booking", 0),
            total.get("order_booking", 0),
            total.get("order_booking_target", 0),
            total.get("order_booking_var_pct"),
        ],
        [
            "Revenue / Billing",
            trading.get("revenue", 0),
            solutions.get("revenue", 0),
            total.get("revenue", 0),
            total.get("revenue_target", 0),
            total.get("revenue_var_pct"),
        ],
        [
            "Gross Margin (₹)",
            trading.get("gross_margin", 0),
            solutions.get("gross_margin", 0),
            total.get("gross_margin", 0),
            total.get("gross_margin_target", 0),
            total.get("gross_margin_var_pct"),
        ],
        [
            "Gross Margin %",
            trading.get("gross_margin_pct"),
            solutions.get("gross_margin_pct"),
            total.get("gross_margin_pct"),
            "",
            "",
        ],
        [
            "No. of Deals Won",
            trading.get("deals_won", 0),
            solutions.get("deals_won", 0),
            total.get("deals_won", 0),
            total.get("deals_won_target", 0),
            "",
        ],
        [
            "Avg Deal Size (₹)",
            trading.get("avg_deal_size", 0),
            solutions.get("avg_deal_size", 0),
            total.get("avg_deal_size", 0),
            "",
            "",
        ],
    ]
    row = _write_table(ws, row, perf_headers, perf_rows)

    row = _write_section_title(ws, row, "2. Top Customers (by Revenue)")
    cust_rows = [
        [
            item.get("customer"),
            item.get("segment_display"),
            item.get("revenue"),
            item.get("gross_margin"),
            item.get("gross_margin_pct"),
            item.get("collection_status"),
        ]
        for item in report.get("top_customers_by_revenue") or []
    ]
    row = _write_table(
        ws,
        row,
        ["Customer", "Segment", "Revenue (₹)", "Gross Margin (₹)", "GM %", "Collection"],
        cust_rows,
    )

    row = _write_section_title(ws, row, "3. Sales Pipeline (forward-looking)")
    pipeline = report.get("forward_pipeline") or {}
    pipe_rows = [
        [
            item.get("customer"),
            item.get("segment_display"),
            item.get("stage"),
            item.get("value"),
            item.get("win_probability"),
            item.get("weighted_value"),
            item.get("expected_close_month"),
        ]
        for item in pipeline.get("opportunities") or []
    ]
    row = _write_table(
        ws,
        row,
        [
            "Opportunity",
            "Segment",
            "Stage",
            "Value (₹)",
            "Win Prob %",
            "Weighted (₹)",
            "Exp. Close",
        ],
        pipe_rows,
    )
    ws.cell(row=row, column=1, value="Total Pipeline Value")
    ws.cell(row=row, column=2, value=pipeline.get("total_pipeline_value", 0))
    row += 1
    ws.cell(row=row, column=1, value="Total Weighted Pipeline")
    ws.cell(row=row, column=2, value=pipeline.get("total_weighted_pipeline", 0))
    row += 2

    row = _write_section_title(ws, row, "4. Lost / Slipped Deals")
    lost_rows = [
        [
            item.get("customer"),
            item.get("value"),
            item.get("stage_lost"),
            item.get("reason"),
            item.get("competitor"),
            item.get("recovery_action"),
        ]
        for item in report.get("lost_deals") or []
    ]
    row = _write_table(
        ws,
        row,
        ["Customer", "Value (₹)", "Stage Lost", "Reason", "Competitor", "Recovery Action"],
        lost_rows,
    )

    row = _write_section_title(ws, row, "Ops Snapshot (CRM)")
    summary_rows = [
        ("Total Leads (period)", summary["total_leads"]),
        ("Open Pipeline Leads (snapshot)", summary["active_pipeline_leads"]),
        ("Won Deals (period)", summary["won_deals"]),
        ("Lost Deals (period)", summary["lost_deals"]),
        ("Win Rate (%) (period)", summary.get("win_rate", 0)),
    ]
    for label, value in summary_rows:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1
    row += 1

    row = _write_section_title(ws, row, "Sales Pipeline by Stage")
    stage_data = [
        [item["stage"], item["count"], item.get("percentage", 0)]
        for item in report["pipeline_by_stage"]
    ]
    row = _write_table(
        ws,
        row,
        ["Stage", "Count", "Percentage (%)"],
        stage_data,
    )

    row = _write_section_title(ws, row, "Category Analysis")
    category_data = [
        [
            item["category"],
            item["lead_count"],
            item["product_quantity"],
            item["pipeline_share_percentage"],
        ]
        for item in report.get("category_analysis", [])
    ]
    row = _write_table(
        ws,
        row,
        ["Category", "Lead Count", "Product Quantity", "Pipeline Share (%)"],
        category_data,
    )

    row = _write_section_title(ws, row, "Salesperson Performance")
    sp_data = [
        [
            item["user"],
            item.get("assigned_leads", item["leads_managed"]),
            item["won_deals"],
            item["lost_deals"],
            item.get("pipeline_product_quantity", 0),
            item.get("win_rate", item["conversion_rate"]),
            item.get("followups_completed", 0),
        ]
        for item in report["salesperson_performance"]
    ]
    row = _write_table(
        ws,
        row,
        [
            "User",
            "Leads Managed (snapshot)",
            "Won (period)",
            "Lost (period)",
            "Pipeline Qty (snapshot)",
            "Win Rate (%)",
            "Follow-ups Completed (period)",
        ],
        sp_data,
    )

    followups = report.get("follow_up_analysis", {})
    if followups:
        row = _write_section_title(ws, row, "Follow-up Analysis")
        followup_rows = [
            ("Today's Follow-ups (live)", followups.get("today_followups", 0)),
            ("Overdue Follow-ups (live)", followups.get("overdue_followups", 0)),
            ("Completed Follow-ups (period)", followups.get("completed_followups", 0)),
        ]
        for label, value in followup_rows:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1
        row += 1

    row = _write_section_title(ws, row, "Top Projects")
    customer_data = [
        [
            item["customer"],
            item["company"],
            item["product_quantity"],
            item["stage"],
        ]
        for item in report["top_customers"]
    ]
    row = _write_table(
        ws,
        row,
        ["Project", "Company", "Product Quantity", "Stage"],
        customer_data,
    )

    products = report.get("products", {})
    if products:
        row = _write_section_title(ws, row, "Open Pipeline — Quantity by Product")
        qty_data = [
            [
                item["product"],
                item.get("category") or "—",
                item.get("brand") or "—",
                item["quantity"],
            ]
            for item in products.get("quantity_by_product", [])
        ]
        row = _write_table(
            ws,
            row,
            ["Product", "Category", "Brand", "Quantity"],
            qty_data,
        )

        row = _write_section_title(ws, row, "Top Selling Products (won in period)")
        sold_data = [
            [
                item["product"],
                item.get("brand") or "—",
                item["quantity"],
            ]
            for item in products.get("top_selling_products", [])
        ]
        row = _write_table(
            ws,
            row,
            ["Product", "Brand", "Quantity"],
            sold_data,
        )

    _auto_width(ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
