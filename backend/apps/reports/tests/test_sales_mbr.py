"""
Sales MBR report API tests — metrics and RBAC.
"""
from datetime import timedelta
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APIClient

from apps.accounts.choices import UserRole
from apps.activities.models import ActivityType, LeadActivity
from apps.leads.models import Brand, Lead, LeadItem, LeadStage, Product, ProductCategory, ProductModel

User = get_user_model()


class SalesMBRReportTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.category = ProductCategory.objects.get(name="IT")
        cls.product = Product.objects.get(category=cls.category, name="Laptop")
        cls.brand = Brand.objects.get(product=cls.product, name="Dell")
        cls.product_model = ProductModel.objects.get(brand=cls.brand, name="Latitude 5540")
        cls.stage_new = LeadStage.objects.get(name="New")
        cls.stage_pre_bid = LeadStage.objects.get(name="Pre Bid")
        cls.stage_won = LeadStage.objects.get(name="Won")
        cls.stage_lost = LeadStage.objects.get(name="Lost")

        cls.ceo = User.objects.create_user(
            username="ceo_mbr",
            email="ceo_mbr@test.com",
            password="pass12345",
            role=UserRole.CEO,
        )
        cls.sales_head = User.objects.create_user(
            username="head_mbr",
            email="head_mbr@test.com",
            password="pass12345",
            role=UserRole.SALES_HEAD,
            manager=cls.ceo,
        )
        cls.sales_head_other = User.objects.create_user(
            username="head_other_mbr",
            email="head_other_mbr@test.com",
            password="pass12345",
            role=UserRole.SALES_HEAD,
            manager=cls.ceo,
        )
        cls.salesperson = User.objects.create_user(
            username="sales_mbr",
            email="sales_mbr@test.com",
            password="pass12345",
            first_name="Raj",
            last_name="Kumar",
            role=UserRole.SALESPERSON,
            manager=cls.sales_head,
        )
        cls.salesperson_other = User.objects.create_user(
            username="sales_other_mbr",
            email="sales_other_mbr@test.com",
            password="pass12345",
            first_name="Other",
            last_name="Rep",
            role=UserRole.SALESPERSON,
            manager=cls.sales_head_other,
        )

        now = timezone.now()
        cls.lead_open = Lead.objects.create(
            customer_name="Acme Corp",
            company_name="Acme Industries",
            category=cls.category,
            stage=cls.stage_pre_bid,
            assigned_to=cls.salesperson,
        )
        Lead.objects.filter(pk=cls.lead_open.pk).update(created_at=now)
        LeadItem.objects.create(
            lead=cls.lead_open,
            category=cls.category,
            product=cls.product,
            brand=cls.brand,
            product_model=cls.product_model,
            quantity=10,
        )

        cls.lead_won = Lead.objects.create(
            customer_name="Beta Ltd",
            company_name="Beta Systems",
            category=cls.category,
            stage=cls.stage_won,
            assigned_to=cls.salesperson,
        )
        Lead.objects.filter(pk=cls.lead_won.pk).update(created_at=now, updated_at=now)
        LeadItem.objects.create(
            lead=cls.lead_won,
            category=cls.category,
            product=cls.product,
            brand=cls.brand,
            product_model=cls.product_model,
            quantity=5,
        )
        LeadActivity.objects.create(
            lead=cls.lead_won,
            user=cls.salesperson,
            activity_type=ActivityType.LEAD_CLOSED_WON,
            comments="Closed won",
        )

        cls.lead_ceo = Lead.objects.create(
            customer_name="CEO Account",
            company_name="Executive Co",
            category=cls.category,
            stage=cls.stage_pre_bid,
            assigned_to=cls.ceo,
        )
        Lead.objects.filter(pk=cls.lead_ceo.pk).update(created_at=now)

        # Older open lead (previous month) — counts in snapshot, not period created.
        old = now - timedelta(days=40)
        cls.lead_old_open = Lead.objects.create(
            customer_name="Legacy Project",
            company_name="Legacy Co",
            category=cls.category,
            stage=cls.stage_pre_bid,
            assigned_to=cls.salesperson,
        )
        Lead.objects.filter(pk=cls.lead_old_open.pk).update(created_at=old, updated_at=old)
        LeadItem.objects.create(
            lead=cls.lead_old_open,
            category=cls.category,
            product=cls.product,
            brand=cls.brand,
            product_model=cls.product_model,
            quantity=3,
        )

    def setUp(self):
        self.client = APIClient()
        self.today = timezone.localdate()
        self.params = {
            "year": self.today.year,
            "month": self.today.month,
        }

    def _auth(self, user):
        self.client.force_authenticate(user=user)

    def test_ceo_can_view_sales_mbr(self):
        self._auth(self.ceo)
        response = self.client.get("/api/v1/reports/sales/", self.params)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("performance_summary", response.data)
        self.assertIn("metric_scopes", response.data)
        self.assertIn("active_pipeline_leads", response.data["performance_summary"])
        self.assertIn("win_rate", response.data["performance_summary"])
        self.assertIn("pipeline_by_stage", response.data)
        self.assertGreaterEqual(
            response.data["performance_summary"]["total_leads"],
            2,
        )

    def test_sales_head_can_view_sales_mbr(self):
        self._auth(self.sales_head)
        response = self.client.get("/api/v1/reports/sales/", self.params)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_sales_head_excludes_ceo_owned_leads(self):
        self._auth(self.sales_head)
        response = self.client.get("/api/v1/reports/sales/", self.params)
        top_customers = response.data["top_customers"]
        companies = [item["company"] for item in top_customers]
        self.assertNotIn("Executive Co", companies)

    def test_salesperson_cannot_access_sales_mbr(self):
        self._auth(self.salesperson)
        response = self.client.get("/api/v1/reports/sales/", self.params)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_unauthenticated_cannot_access_sales_mbr(self):
        response = self.client.get("/api/v1/reports/sales/", self.params)
        self.assertIn(
            response.status_code,
            [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN],
        )

    def test_salesperson_filter(self):
        self._auth(self.ceo)
        response = self.client.get(
            "/api/v1/reports/sales/",
            {**self.params, "salesperson": str(self.salesperson.id)},
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["filters"]["salesperson_id"], str(self.salesperson.id))

    def test_ceo_assignee_available_in_filter(self):
        self._auth(self.ceo)
        response = self.client.get("/api/v1/reports/sales/", self.params)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        assignee_ids = [item["id"] for item in response.data["salespeople"]]
        self.assertIn(str(self.ceo.id), assignee_ids)
        ceo_option = next(
            item for item in response.data["salespeople"] if item["id"] == str(self.ceo.id)
        )
        self.assertIn("(CEO)", ceo_option["name"])

    def test_ceo_assignee_filter_shows_ceo_owned_leads(self):
        self._auth(self.ceo)
        response = self.client.get(
            "/api/v1/reports/sales/",
            {**self.params, "salesperson": str(self.ceo.id)},
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        companies = [item["company"] for item in response.data["top_customers"]]
        self.assertIn("Executive Co", companies)

    def test_sales_head_assignee_filter_is_team_only(self):
        self._auth(self.sales_head)
        response = self.client.get("/api/v1/reports/sales/", self.params)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        assignee_ids = {item["id"] for item in response.data["salespeople"]}
        self.assertIn(str(self.sales_head.id), assignee_ids)
        self.assertIn(str(self.salesperson.id), assignee_ids)
        self.assertNotIn(str(self.salesperson_other.id), assignee_ids)
        self.assertNotIn(str(self.ceo.id), assignee_ids)

    def test_sales_head_cannot_filter_outside_team(self):
        self._auth(self.sales_head)
        response = self.client.get(
            "/api/v1/reports/sales/",
            {**self.params, "salesperson": str(self.salesperson_other.id)},
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_invalid_month_returns_400(self):
        self._auth(self.ceo)
        response = self.client.get(
            "/api/v1/reports/sales/",
            {"year": self.today.year, "month": 13},
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_ceo_can_export_excel(self):
        self._auth(self.ceo)
        response = self.client.get("/api/v1/reports/sales/export/", self.params)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn(
            "spreadsheetml",
            response["Content-Type"],
        )
        workbook = load_workbook(BytesIO(b"".join(response.streaming_content)))
        self.assertIn("Sales MBR", workbook.sheetnames)
        ws = workbook["Sales MBR"]
        self.assertEqual(ws["A1"].value.split("—")[0].strip(), "Sales MBR")

    def test_sales_head_can_export_excel(self):
        self._auth(self.sales_head)
        response = self.client.get("/api/v1/reports/sales/export/", self.params)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_salesperson_cannot_export_excel(self):
        self._auth(self.salesperson)
        response = self.client.get("/api/v1/reports/sales/export/", self.params)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_won_deals_counted_in_summary(self):
        self._auth(self.ceo)
        response = self.client.get("/api/v1/reports/sales/", self.params)
        self.assertGreaterEqual(response.data["performance_summary"]["won_deals"], 1)
        self.assertGreaterEqual(
            response.data["performance_summary"]["won_product_quantity"],
            5,
        )

    def test_salesperson_performance_includes_assignee(self):
        self._auth(self.ceo)
        response = self.client.get("/api/v1/reports/sales/", self.params)
        names = [row["user"] for row in response.data["salesperson_performance"]]
        self.assertIn("Raj Kumar", names)

    def test_open_pipeline_snapshot_includes_prior_month_leads(self):
        self._auth(self.ceo)
        response = self.client.get("/api/v1/reports/sales/", self.params)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        summary = response.data["performance_summary"]
        # Snapshot open pipeline includes current + prior-month open leads (+ CEO).
        self.assertGreaterEqual(summary["active_pipeline_leads"], 3)
        self.assertGreaterEqual(summary["pipeline_product_quantity"], 13)
        # Period created excludes the 40-day-old lead.
        self.assertGreaterEqual(summary["total_leads"], 3)
        companies = [item["company"] for item in response.data["top_customers"]]
        self.assertIn("Legacy Co", companies)

    def test_prior_month_excludes_current_period_created_leads(self):
        self._auth(self.ceo)
        prior = self.today.replace(day=1) - timedelta(days=1)
        response = self.client.get(
            "/api/v1/reports/sales/",
            {"year": prior.year, "month": prior.month},
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        summary = response.data["performance_summary"]
        # Old lead was created ~40 days ago; may or may not fall in prior calendar month.
        # Current-month won deals must not appear in a different period.
        self.assertEqual(summary["won_deals"], 0)
        self.assertEqual(summary["won_product_quantity"], 0)
        # Snapshot still shows live open pipeline.
        self.assertGreaterEqual(summary["active_pipeline_leads"], 3)

    def test_products_respect_assignee_filter(self):
        self._auth(self.ceo)
        response = self.client.get(
            "/api/v1/reports/sales/",
            {**self.params, "salesperson": str(self.ceo.id)},
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        products = response.data["products"]["quantity_by_product"]
        # CEO open lead has no line items in fixture → empty pipeline products.
        self.assertEqual(products, [])
        self.assertEqual(response.data["products"]["top_selling_products"], [])

    def test_top_customers_include_lead_id(self):
        self._auth(self.ceo)
        response = self.client.get("/api/v1/reports/sales/", self.params)
        self.assertTrue(response.data["top_customers"])
        self.assertIn("lead_id", response.data["top_customers"][0])

    def test_sales_pack_sections_present(self):
        self._auth(self.ceo)
        response = self.client.get("/api/v1/reports/sales/", self.params)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("sales_performance", response.data)
        self.assertIn("trading", response.data["sales_performance"])
        self.assertIn("solutions", response.data["sales_performance"])
        self.assertIn("total", response.data["sales_performance"])
        self.assertIn("forward_pipeline", response.data)
        self.assertIn("total_weighted_pipeline", response.data["forward_pipeline"])
        self.assertIn("top_customers_by_revenue", response.data)
        self.assertIn("lost_deals", response.data)

    def test_won_deal_commercial_metrics(self):
        self.lead_won.deal_value = 100000
        self.lead_won.billed_amount = 100000
        self.lead_won.gross_margin_amount = 20000
        self.lead_won.business_segment = "TRADING"
        self.lead_won.save(
            update_fields=[
                "deal_value",
                "billed_amount",
                "gross_margin_amount",
                "business_segment",
                "updated_at",
            ],
        )
        self._auth(self.ceo)
        response = self.client.get("/api/v1/reports/sales/", self.params)
        total = response.data["sales_performance"]["total"]
        self.assertGreaterEqual(total["order_booking"], 100000)
        self.assertGreaterEqual(total["revenue"], 100000)
        self.assertGreaterEqual(total["gross_margin"], 20000)
        self.assertGreaterEqual(total["deals_won"], 1)
